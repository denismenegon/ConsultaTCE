import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import Calendar
import pandas as pd
from playwright.sync_api import sync_playwright
from datetime import datetime
import time
import os
import sys
import re
import openpyxl
import threading
from openpyxl import load_workbook
from unidecode import unidecode

intLinha = 0


# Defina o caminho correto para os navegadores dentro do executável
if getattr(sys, 'frozen', False):
    # Quando o código está sendo executado a partir de um executável
    caminho_playwright = os.path.join(sys._MEIPASS, 'ms-playwright')
else:
    # Quando o código está sendo executado a partir do ambiente de desenvolvimento
    # caminho_playwright = 'C:/Users/denis.menegon/AppData/Local/ms-playwright'

    caminho_playwright = 'C:/Users/Denis Menegon/AppData/Local/ms-playwright'

# Defina o caminho para os navegadores
os.environ['PLAYWRIGHT_BROWSERS_PATH'] = caminho_playwright

# Agora, inicie o Playwright normalmente
from playwright.sync_api import sync_playwright

with sync_playwright() as p:
    # navegador = p.chromium.launch(headless=False)

    #   navegador =  p.edge.launch(headless=False)  

    p.firefox.launch() 


def selecionar_arquivo():
    # Abre a janela de seleção de arquivo
    arquivo = filedialog.askopenfilename(title="Selecione um arquivo", filetypes=[("Arquivos Excel", "*.xlsx"), ("Arquivos de texto", "*.txt")])
    if arquivo:
        # Exibe o caminho do arquivo no campo de texto (somente leitura)
        caminho_arquivo_var.set(arquivo)

def gravar_informacao(nome_arquivo, mensagem):
    with open(nome_arquivo, 'a') as file:
        file.write(f"{mensagem}\n")

def processar_arquivo():
    i = 0
    abaAtual = 0
    linhaAba = 1

    arquivo = caminho_arquivo_var.get()
    
    # Se a barra de progresso já existir, remova-a
    global progress_bar  # Tornando a progress_bar uma variável global
    

    if not arquivo:
        messagebox.showerror("Erro", "Nenhum arquivo selecionado!")
        return
    
    try:

        # Identifica o tipo de arquivo
        if arquivo.endswith(".txt"):
            dados = ler_arquivo_txt(arquivo)
        elif arquivo.endswith(".xlsx"):
            dados = ler_arquivo_excel(arquivo,  combobox_entidades_var.get().strip())
        else:
            raise ValueError("Formato de arquivo não suportado.")
        

        global intLinha 

        if progress_bar is not None:
            progress_bar.grid_forget()

        # Cria uma nova barra de progresso com o novo comprimento
        progress_bar = None 
        

        # Verifica o valor selecionado no primeiro combobox e ajusta os valores do segundo
        if combobox_documento_parametro.get() == "Visualizar Documentos Públicos":
            progress_bar = ttk.Progressbar(root, orient="horizontal", length=20, mode="determinate", maximum=intLinha+0.5)

            progress_bar.grid(row=13, column=2, padx=5, pady=5, sticky="ew")
            label_progresso.grid(row=13, column=0, columnspan=3, padx=5, pady=5, sticky="w")

        else:
            progress_bar = ttk.Progressbar(root, orient="horizontal", length=117, mode="determinate", maximum=intLinha+0.5)
            progress_bar.grid(row=10, column=2, padx=5, pady=5, sticky="w")

            label_progresso.grid(row=10, column=0, columnspan=3, padx=5, pady=10, sticky="w")  


        label_progresso.config(text=f"Executando...")  # Atualiza o texto do Label

        progress_bar['value'] = 1

        print (intLinha)

        regraProcessamento(blnProcessar=True)

        # Captura o link da tela (não vamos usar filtros por enquanto)
        link_tela = link_var.get()  # Link informado na tela
        
        # Se o campo do link na tela estiver preenchido, utiliza esse link
        if link_tela:
            print(f"Usando o link informado na tela: {link_tela}")
            dados = [(usuario, senha, link_tela, indice_aba, cidade) for usuario, senha, link, indice_aba, cidade in dados]
        
        # Verifica os dados lidos
        if dados:
            print(f"Dados lidos: {dados}")  # Verificando os dados lidos
            for usuario, senha, link, indice_aba, cidade  in dados:
                print(f"Processando usuário: {usuario}, senha: {senha}, link: {link}, Aba Corrente: {indice_aba}")  # Verificação de dados
                
                i += 1
                linhaAba += 1

                # Transição de Aba
                if abaAtual != indice_aba:
                    linhaAba=2

                label_progresso.config(text=f"{cidade} - {usuario} - {i} de {intLinha}")  # Atualiza o texto do Label
                
                print(f"Linha da Planilha: {i}")

                # Obtenha o valor do combobox aqui
                menu = combobox_menu_var.get().strip()  
                perfil = combobox_perfil_var.get().strip()
                parametroAno = ano_var.get()
                parametroMesInicial = mes_inicial_var.get()
                parametroMesFinal = mes_final_var.get()
                parametroCriterioPesquisa = combobox_criterio_pesquisa_var.get().strip()  # Obtém o valor do combobox e remove espaços extras
                parametroDocumento = combobox_documento_var.get().strip()
                parametroDtInicial = entry_data_inicial.get()
                parametroDtFinal = entry_data_final.get()
                parametroMesReferencia = combobox_mes_referencia.get()
                parametroTipoProcesso = combobox_tipo_processo.get()

                parametroCidadeCorrente = cidade
                usuarioPlanilha =  str(usuario).strip().lower() 
                senhaPlanilha =  str(senha).strip().lower() 

                processar_com_playwright(usuario, senha, link, menu, perfil, parametroAno, parametroMesInicial, parametroMesFinal, parametroCriterioPesquisa, parametroDocumento, parametroDtInicial, parametroDtFinal, parametroMesReferencia, parametroTipoProcesso, indice_aba, parametroCidadeCorrente, linhaAba)  # Passe o parâmetro

                abaAtual = indice_aba

                progress_bar['value'] = i + 1
                root.update_idletasks()  # Atualiza a interface
            messagebox.showinfo("Sucesso", "Arquivo processado com sucesso!")

        else:
            messagebox.showerror("Erro", "Não foram encontrados dados válidos.")
            
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao processar o arquivo: {str(e)}")

        label_progresso.config(text="")

        regraProcessamento(blnProcessar=False)

        progress_bar.grid_forget()

    label_progresso.config(text="")

    regraProcessamento(blnProcessar=False)

    progress_bar.grid_forget()


def regraProcessamento(blnProcessar):
    if blnProcessar:
        entry_data_inicial.config(state="disabled")
        entry_data_final.config(state="disabled")
        combobox_criterio_pesquisa_parametro.config(state="disabled")
        combobox_tipo_processo.config(state="disabled")
        combobox_documento_parametro.config(state="disabled")
        combobox_perfil_parametro.config(state="disabled")
        combobox_menu_parametro.config(state="disabled")
        checkbox_background.config(state="disabled")
        combobox_mes_inicial.config(state="disabled")
        combobox_mes_fim.config(state="disabled")
        ano_combobox.config(state="disabled")
        combobox_mes_referencia.config(state="disabled")
        button_processar.config(state="disabled") 
        entrada_url.config(state="disabled") 
        button_selecionar_arquivo.config(state="disabled")
        combobox_entidades.config(state="disabled") 
    else:
        entry_data_inicial.config(state="normal")
        entry_data_final.config(state="normal")
        combobox_criterio_pesquisa_parametro.config(state="readonly")
        combobox_tipo_processo.config(state="readonly")
        combobox_documento_parametro.config(state="readonly")
        combobox_perfil_parametro.config(state="readonly")
        combobox_menu_parametro.config(state="disabled")
        checkbox_background.config(state="normal")
        combobox_mes_inicial.config(state="readonly")
        combobox_mes_fim.config(state="readonly")
        ano_combobox.config(state="readonly")
        combobox_mes_referencia.config(state="readonly")
        button_processar.config(state="normal")  
        entrada_url.config(state="normal") 
        button_selecionar_arquivo.config(state="normal")
        combobox_entidades.config(state="readonly")

def ler_arquivo_txt(arquivo):
    # Lê o arquivo .txt e extrai os dados
    dados = []
    try:
        with open(arquivo, 'r') as file:
            for linha in file.readlines():
                # Imprime cada linha lida para depuração
                print(f"Linha lida do arquivo TXT: {linha.strip()}")
                # Assume que o formato da linha é: usuario, senha, link
                partes = linha.strip().split(',')
                if len(partes) == 3:
                    dados.append(partes)
                else:
                    print(f"Formato incorreto na linha: {linha.strip()}")
    except Exception as e:
        print(f"Erro ao ler arquivo TXT: {e}")
    return dados


def ler_arquivo_excel(arquivo, aba_especifica="Todos"):
    # Lê o arquivo .xlsx e extrai os dados
    dados = []

    try:
        # Lê todas as abas do arquivo Excel
        planilhas = pd.read_excel(arquivo, sheet_name=None)  # Lê todas as abas (sheet_name=None)
        
        global intLinha
        intLinha = 0
        

        # Se o parâmetro "aba_especifica" não for "Todos", verifica se a aba existe no arquivo
        if aba_especifica != "Todos" and aba_especifica != "Outros" and aba_especifica not in planilhas:
            print(f"A aba '{aba_especifica}' não foi encontrada no arquivo.")
            return dados  # Retorna a lista vazia se a aba não existir
        

        # Condicional para ler "Todos", "Prefeitura", "Câmera" ou "Outros"
        if aba_especifica == "Todos":
            abas_a_ler = [(i, nome_aba, df) for  i, (nome_aba, df) in enumerate(planilhas.items())]  # Lê todas as abas
        elif aba_especifica == "Outros":
            # Exclui as abas "Prefeitura" e "Câmera" e processa as restantes
            abas_a_ler = [(i, nome_aba, df) for  i, (nome_aba, df) in enumerate(planilhas.items()) if nome_aba not in ["Prefeitura", "Câmara"]]
        else:
            # Lê apenas a aba especificada
            abas_a_ler = [(i, nome_aba, df) for  i, (nome_aba, df) in enumerate(planilhas.items()) if nome_aba in [aba_especifica]]



        print(f"Aba(s) encontrada(s): {list(planilhas.keys())}")


        # Itera sobre todas as abas e suas respectivas planilhas
        # for iAba, (nome_aba, df) in enumerate(planilhas.items()):
        for iAba, nome_aba, df in abas_a_ler:
            # print(f"Dados da aba {nome_aba}: {df.head()}")  # Mostra os primeiros registros de cada aba para depuração

            # df = pd.read_excel(arquivo)
            # print(f"Dados lidos do arquivo Excel: {df.head()}")  # Mostra os primeiros registros para depuração
            
  
            # Ajuste para pegar os dados de Login e Senha
            for index, row in df.iterrows():
                # Ajuste para pegar o Login e a Senha corretamente (adapte conforme os nomes exatos das colunas)
                usuario = row.get('Login')  # Ajuste aqui se a coluna tiver outro nome
                senha = row.get('Senha')    # Ajuste aqui se a coluna tiver outro nome
                link = link_var.get()      # Link da tela (se informado)
                cidade =  row.get('Entidade')

                if usuario and senha:
                    # Se houver link na tela, use o link informado, caso contrário, use o link da planilha
                    link = link if link else row.get('Link')  # Ajuste aqui se a coluna "Link" existir
                    if link:
                        dados.append((usuario, senha, link, iAba, cidade))

                          # Informa ao Python que estamos usando a variável global
                        intLinha += 1
                    else:
                        print(f"Link não encontrado na planilha ou na tela para o usuário {usuario}.")
                else:
                    print(f"Login ou Senha não encontrados na linha {index}")

    except Exception as e:
        print(f"Erro ao ler arquivo Excel: {e}")
    return dados


# def wait_for_element_with_retry(pagina, selector, timeout=30000, retries=3, interval=5):
def wait_for_element_with_retry(pagina, selector, timeout=30000, retries=3, interval=5):
    """Função para aguardar um elemento com retries em caso de lentidão"""
    attempt = 0
    while attempt < retries:
        try:
            # Tenta encontrar o elemento
            pagina.wait_for_selector(selector, timeout=timeout)
            return True
        except Exception as e:
            attempt += 1
            print(f"Tentativa {attempt} de aguardar o elemento {selector}. Erro: {e}")
            if attempt >= retries:
                return False
            time.sleep(interval)  # Espera antes de tentar novamente
    return False


# Função para limpar e normalizar a string (remover acentos e espaços extras)
def normalize_text(text):
    return unidecode(str(text).strip().lower())  # Remove acentos e converte para minúsculas, mantendo os espaços


# Função para verificar o index do item no select
def obter_index_combobox(page, nome_item, selector):
    # Seleciona o elemento select
    select_element = page.locator(selector)
    
    # Obtém todas as opções (itens) dentro do select
    options = select_element.locator('option')

    # Varrendo as opções e verificando o nome
    count = options.count()  # Quantidade de opções dentro do select
    for index in range(count):
        option_text = options.nth(index).text_content()

        # Normalização e remoção de espaços
        option_text_normalized = normalize_text(option_text)
        nome_item_normalized = normalize_text(nome_item)

        # if unidecode(option_text.strip()).lower().startswith(unidecode(nome_item).lower()):


        print(f"option_text_normalized: {repr(option_text_normalized)}")
        print(f"nome_item_normalized: {repr(nome_item_normalized)}")



        if option_text_normalized in nome_item_normalized:
            return index  # Retorna o índice do item encontrado
    return -1  # Retorna -1 caso o item não seja encontrado


def processar_com_playwright(usuario, senha, link, menu, perfil, ano, mesInicial, mesFinal, criterioPesquisa, documento, dtInicial = "", dtFinal = "", mesReferencia = "", tipoProcesso = "", abaCorrente = 0, cidadeCorrente = "", linhaCorrente = 0, blnUnicoMunicipio = True):
    try:
        with sync_playwright() as p:
            navegador = p.chromium.launch(headless = True if checkbox_background_var.get() == 1  else False)
            contexto = navegador.new_context()
            pagina = contexto.new_page()


            # blnUnicoMunicipio = False
            

            # Adiciona o evento para verificar erros 404
            pagina.on("response", lambda response: print(f"Status code: {response.status} URL: {response.url}"))

            pagina.goto(link)
            print(f"[{datetime.now()}] Página carregada: {link}")

            # Garantir que os valores de usuário e senha são strings
            usuario = str(usuario).strip()  # Converte para string e remove espaços extras
            senha = str(senha).strip()      # Converte para string e remove espaços extras

            # Preenche os campos de login
            pagina.fill('input[id="username"]', usuario)
            pagina.fill('input[id="password"]', senha)
            pagina.click('input[name="submit"]')
            time.sleep(5)  # Espera para garantir que a página pós-login tenha carregado

            # Localiza o elemento com o texto exato
            login = pagina.locator("span:text('Usuário não encontrado ou senha incorreta.')")
            
            # Verifica se o elemento está visível na página
            if login.is_visible():
                atualizar_planilha(usuario, 'Sem Acesso', 'G', abaCorrente, linhaCorrente)
                atualizar_planilha(usuario, '', 'H', abaCorrente, linhaCorrente)

                return

            link_principal = pagina.url

            # Definir o menu baseado no parâmetro
            if menu.lower() == "default01":
                strMenu = 'label:has-text("Auditoria Eletrônica de Órgãos Públicos")'
            elif menu.lower() == "default02":
                strMenu = 'label:has-text("Outro Menu")'
            else:
                strMenu = 'label:has-text("Auditoria Eletrônica de Órgãos Públicos")'

            if not wait_for_element_with_retry(pagina, strMenu):
                print(f"[{datetime.now()}] Erro: Menu não encontrado.")
                return

            with contexto.expect_page() as nova_pagina_info:
                pagina.click(strMenu)
                
            nova_pagina = nova_pagina_info.value
            print(f"[{datetime.now()}] Menu clicado.")

            # Aguarda o carregamento por completo
            nova_pagina.wait_for_load_state('load') 

            
            if not wait_for_element_with_retry(nova_pagina, 'select#perfil-usuario', 1000, 1, 1):
                print(f"[{datetime.now()}] Erro: Combobox de perfil não carregado.")
                
                if wait_for_element_with_retry(nova_pagina, 'li:has-text("Documento") > a', 1000, 2, 1):
                    print("Menu Documento existe!")

                    time.sleep(1)

                    nova_pagina.click('li:has-text("Documento") > a')
            else:
                nova_pagina.select_option('select#perfil-usuario', label="Audesp Base - Prestação de dados")
                time.sleep(2)

                nova_pagina.click('button#button-ok')
                time.sleep(2)

                nova_pagina.click('li:has-text("Documento") > a')
                time.sleep(2)

            # Visualizar Documentos Públicos
            if documento == "Visualizar Documentos Públicos":
                vtrDtRecebimento = []
                vtrStatus = []

                nova_pagina.click('a[href="./visualizarDocumentoPublico.do"]')

                time.sleep(2)

                nova_pagina.wait_for_load_state('networkidle')
                posicao_dados_url = nova_pagina.url

                # Espera os filtros estarem visíveis
                if not wait_for_element_with_retry(nova_pagina, 'select[name="mesReferencia"]'):
                    print(f"[{datetime.now()}] Erro: Filtro 'Mes Referência' não carregado.")
                    return

                municipio_options = nova_pagina.query_selector_all('select[name="municipio"] option')
                municipio_nome = nova_pagina.locator('select[name="municipio"] option:checked').text_content()

                time.sleep(1)

                print(f"Quantidade de Itens do Município: {len(municipio_options)}")

                # Seleção do Município
                if len(municipio_options) > 1:                   
                    # for municipio_index in range(1, len(municipio_options)):  # Começar com o índice 1
                    for municipio_index in range(1 if not blnUnicoMunicipio else 0, len(municipio_options) if not blnUnicoMunicipio else 1):
                    
                        time.sleep(1)

                        nova_pagina.select_option('select[name="mesReferencia"]', value=mesReferencia)

                        # Preencher o campo de input com o valor da variável
                        nova_pagina.fill('input[name="exercicio"]', ano)  # Preenche o campo com o valor de 'ano'

                        nova_pagina.select_option('select[name="tipoDocumento"]', label=criterioPesquisa)

                        nova_pagina.fill('input[name="dataInicialReceb"]', dtInicial)  # Preenche o campo com o valor de 'ano'
                        nova_pagina.fill('input[name="dataFinalReceb"]', dtFinal)  # Preenche o campo com o valor de 'ano'
                        
                        nova_pagina.select_option('select[name="tpProcessoId"]', label=tipoProcesso)

                        time.sleep(1)

                        
                        if not blnUnicoMunicipio:
                            if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="municipio"]', municipio_index):
                                print(f"[{datetime.now()}] Erro ao selecionar o município {municipio_index}.")
                                continue
                            
                        else:
                            municipio_index = obter_index_combobox(nova_pagina, cidadeCorrente, 'select[name="municipio"]')
                            print(f"Cidade Atual: {cidadeCorrente}")
                            print (f"Índice do Município {municipio_index}")
                            if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="municipio"]', municipio_index):
                                print(f"[{datetime.now()}] Erro ao selecionar o município {municipio_index}.")
                                continue
                        

                        municipio_nome = nova_pagina.locator('select[name="municipio"] option:checked').text_content()

                        time.sleep(1)

                        # Agora pega todas as opções
                        entidade_options = nova_pagina.query_selector_all('select[name="entidade"] option')

                        print(f"################## Município {municipio_index} Quantidade Total para Entidade {len(entidade_options)}. ##################")

                        if len(entidade_options) > 1:
                            for entidade_index in range(1 if not blnUnicoMunicipio else 0, len(entidade_options) if not blnUnicoMunicipio else 1):  # Começar com o índice 1 para Entidade
                                print(f"[{datetime.now()}] Processando Entidade {entidade_index} para o Município {municipio_index}")

                                # Na condição ter mais Item, necessário selecionar o Munícipio corrente para poder carregar o ComboBox da Entidade
                                if entidade_index > 1:
                                    # Selecionar o elemento <select> e, em seguida, escolher a opção pelo índice
                                    # nova_pagina.select_option('select[name="municipio"]', index=municipio_index)
                                    nova_pagina.select_option('select[name="mesReferencia"]', value=mesReferencia)

                                    # Preencher o campo de input com o valor da variável
                                    nova_pagina.fill('input[name="exercicio"]', ano)  # Preenche o campo com o valor de 'ano'

                                    nova_pagina.select_option('select[name="tipoDocumento"]', label=criterioPesquisa)

                                    nova_pagina.fill('input[name="dataInicialReceb"]', dtInicial)  # Preenche o campo com o valor de 'ano'
                                    nova_pagina.fill('input[name="dataFinalReceb"]', dtFinal)  # Preenche o campo com o valor de 'ano'
                                    
                                    nova_pagina.select_option('select[name="tpProcessoId"]', label=tipoProcesso)

                                    time.sleep(1)

                                    if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="municipio"]', municipio_index):
                                        print(f"[{datetime.now()}] Erro ao selecionar o município {municipio_index}.")
                                        continue

                                    time.sleep(1)

                                if blnUnicoMunicipio:
                                    if len(entidade_options) > 2:
                                        entidade_index = obter_index_combobox(nova_pagina, cidadeCorrente, 'select[name="entidade"]')
                                        print(f"Cidade Atual: {cidadeCorrente}")
                                        print (f"Índice da Entidade {municipio_index}")

                                        if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="entidade"]', entidade_index):
                                            print(f"[{datetime.now()}] Erro ao selecionar o município {entidade_index}.")
                                            continue
                               
                                else:
                                    if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="entidade"]', entidade_index):
                                        print(f"[{datetime.now()}] Erro ao selecionar a entidade {entidade_index} para o município {municipio_index}.")
                                        continue


                                # Verificando e clicando no botão 'Pesquisar'
                                if not wait_for_element_with_retry(nova_pagina, 'input[value="Pesquisar"]:visible'):
                                    print(f"[{datetime.now()}] Erro: Botão 'Pesquisar' não visível.")
                                    return
                                
                                # Localizando o botão pelo valor do atributo `value`
                                nova_pagina.click('input[value="Pesquisar"]')
                                                                
                                print(f"[{datetime.now()}] Botão 'Pesquisar' clicado para Município {municipio_index}, Entidade {entidade_index}")

                                time.sleep(2)

                                # Verifica se a tabela existe
                                try:
                                    # Verifique se a tabela contém a mensagem "Não há registros para o critério informado."
                                    tabela = nova_pagina.query_selector('table#item')

                                    # Caso contrário, continue a iteração sobre as linhas da tabela
                                    rows = tabela.query_selector_all('tbody tr')  # Seleciona todas as linhas <tr> dentro do <tbody>
                                    for row in rows:
                                        cols = row.query_selector_all('td')  # Seleciona todas as colunas <td> de cada linha
                                        if cols:
                                            id_pacote = cols[0].inner_text()
                                            metadado = cols[1].inner_text()
                                            tipo = cols[2].inner_text()
                                            recebimento = cols[3].inner_text()
                                            categoria = cols[4].inner_text()

                                            entidade = cols[5].inner_text()
                                            municipio = cols[6].inner_text()
                                            mes = cols[7].inner_text()
                                            exercicio = cols[8].inner_text()
                                            status = cols[9].inner_text()
                                            tipo_processo = cols[10].inner_text()

                                            vtrDtRecebimento.append(recebimento)
                                            vtrStatus.append(status)

                                            # Converte as datas em objetos datetime
                                            datas_convertidas = [converter_data(data) for data in vtrDtRecebimento]

                                            # Encontra o índice e a data mais atual
                                            i = max(enumerate(datas_convertidas), key=lambda x: x[1])[0]

                                            gravar_informacao(f'dados_pacotes_visualizar_documento_publico_{data_hoje}.txt', f'{cidadeCorrente}, {id_pacote}, {metadado}, {tipo}, {recebimento}, {categoria}, {entidade}, {categoria}, {municipio}, {mes}, {exercicio}, {status}, {tipo_processo}')

                                            print(f"Linha : {linhaCorrente}")

                                            atualizar_planilha(usuario, vtrStatus[i], 'G', abaCorrente, linhaCorrente)
                                            atualizar_planilha(usuario, vtrDtRecebimento[i], 'H', abaCorrente, linhaCorrente)

                                            processar_paginacao_por_numeros(nova_pagina)

                                            time.sleep(2)

                                except Exception as e:
                                    gravar_informacao(f'dados_pacotes_visualizar_documento_publico_{data_hoje}.txt', f'{cidadeCorrente}  -  {criterioPesquisa} não encontrado')
                                                      
                                    atualizar_planilha(usuario, f'{criterioPesquisa} não encontrado' , 'G', abaCorrente, linhaCorrente)
                                    atualizar_planilha(usuario, '', 'H', abaCorrente, linhaCorrente)

                                    print(f"[{datetime.now()}] Erro ao verificar a tabela: {e}")

                                # Volta para a URL inicial de "Posição Dados Transmitidos"
                                nova_pagina.goto(posicao_dados_url)
                                time.sleep(2)

                        else:
                            print("Única entidade!!")

                            nova_pagina.select_option('select[name="municipio"]', index=1)
                            #Altera para o primeito Item 
                            nova_pagina.select_option('select[name="entidade"]', index=1)

                            # Verificando e clicando no botão 'Pesquisar'
                            if not wait_for_element_with_retry(nova_pagina, 'input[value="Pesquisar"]:visible'):
                                print(f"[{datetime.now()}] Erro: Botão 'Pesquisar' não visível.")
                                return
                            
                            # Localizando o botão pelo valor do atributo `value`
                            nova_pagina.click('input[value="Pesquisar"]')
                            
                            print(f"[{datetime.now()}] Botão 'Pesquisar' clicado para Município {municipio_index}")

                            time.sleep(2)

                            # Verifica se a tabela existe
                            try:
                                # Verifique se a tabela contém a mensagem "Não há registros para o critério informado."
                                tabela = nova_pagina.query_selector('table#item')

                                # Caso contrário, continue a iteração sobre as linhas da tabela
                                rows = tabela.query_selector_all('tbody tr')  # Seleciona todas as linhas <tr> dentro do <tbody>
                                for row in rows:
                                    cols = row.query_selector_all('td')  # Seleciona todas as colunas <td> de cada linha
                                    if cols:
                                        id_pacote = cols[0].inner_text()
                                        metadado = cols[1].inner_text()
                                        tipo = cols[2].inner_text()
                                        recebimento = cols[3].inner_text()
                                        categoria = cols[4].inner_text()

                                        entidade = cols[5].inner_text()
                                        municipio = cols[6].inner_text()
                                        mes = cols[7].inner_text()
                                        exercicio = cols[8].inner_text()
                                        status = cols[9].inner_text()
                                        tipo_processo = cols[10].inner_text()

                                        vtrDtRecebimento.append(recebimento)
                                        vtrStatus.append(status)

                                        # Converte as datas em objetos datetime
                                        datas_convertidas = [converter_data(data) for data in vtrDtRecebimento]

                                        # Encontra o índice e a data mais atual
                                        i = max(enumerate(datas_convertidas), key=lambda x: x[1])[0]

                                        gravar_informacao(f'dados_pacotes_visualizar_documento_publico_{data_hoje}.txt', f'{cidadeCorrente}, {id_pacote}, {metadado}, {tipo}, {recebimento}, {categoria}, {entidade}, {categoria}, {municipio}, {mes}, {exercicio}, {status}, {tipo_processo}')

                                        atualizar_planilha(usuario, vtrStatus[i], 'G', abaCorrente, linhaCorrente)
                                        atualizar_planilha(usuario, vtrDtRecebimento[i], 'H', abaCorrente, linhaCorrente)
                                        
                                        processar_paginacao_por_numeros(nova_pagina)

                                        time.sleep(2)
                            except Exception as e:
                                gravar_informacao(f'dados_pacotes_visualizar_documento_publico_{data_hoje}.txt', f'{cidadeCorrente}  -  {criterioPesquisa} não encontrado')

                                atualizar_planilha(usuario, f'{criterioPesquisa} não encontrado' , 'G', abaCorrente, linhaCorrente)
                                atualizar_planilha(usuario, '', 'H', abaCorrente, linhaCorrente)

                                print(f"[{datetime.now()}] Erro ao verificar a tabela: {e}")

                                # Volta para a URL inicial de "Posição Dados Transmitidos"
                                nova_pagina.goto(posicao_dados_url)
                                time.sleep(2)
                                
                
                else:
                    print(f"[{datetime.now()}] Somente um Município, verificando Entidades.")
                    entidade_options = nova_pagina.query_selector_all('select[name="entidade"] option')
                    if len(entidade_options) > 1:
                        for entidade_index in range(1, len(entidade_options)):  # Começar com o índice 1 para Entidade
                            print(f"[{datetime.now()}] Processando Entidade {entidade_index} para o único Município")

                            if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="entidade"]', entidade_index):
                                print(f"[{datetime.now()}] Erro ao selecionar a entidade {entidade_index}.")
                                continue

                            # Espera a página carregar e realiza a pesquisa
                            nova_pagina.wait_for_load_state('networkidle', timeout=30000)
                           
                            # Verificando e clicando no botão 'Pesquisar'
                            if not wait_for_element_with_retry(nova_pagina, 'input[value="Pesquisar"]:visible'):
                                print(f"[{datetime.now()}] Erro: Botão 'Pesquisar' não visível.")
                                return
                            
                            # Localizando o botão pelo valor do atributo `value`
                            nova_pagina.click('input[value="Pesquisar"]')

                            print(f"[{datetime.now()}] Botão 'Pesquisar' clicado para Entidade {entidade_index}")

                            # Verifica se a tabela existe
                            try:
                                # Verifique se a tabela contém a mensagem "Não há registros para o critério informado."
                                mensagem_erro = tabela.query_selector('tbody tr.empty')
                                if mensagem_erro:
                                    mensagem_texto = mensagem_erro.inner_text()
                                    if "Não há registros para o critério informado." in mensagem_texto:
                                        print(f"[{datetime.now()}] Não há registros para o critério informado para o Município {municipio_index}, Entidade {entidade_index}")
                                    else:
                                        tabela = nova_pagina.query_selector('table')

                                        # Caso contrário, continue a iteração sobre as linhas da tabela
                                        rows = tabela.query_selector_all('tbody tr')  # Seleciona todas as linhas <tr> dentro do <tbody>

                                        for row in rows:
                                            cols = row.query_selector_all('td')  # Seleciona todas as colunas <td> de cada linha
                                            if cols:
                                                id_pacote = cols[0].inner_text()
                                                metadado = cols[1].inner_text()
                                                tipo = cols[2].inner_text()
                                                recebimento = cols[3].inner_text()
                                                categoria = cols[4].inner_text()

                                                entidade = cols[5].inner_text()
                                                municipio = cols[6].inner_text()
                                                mes = cols[7].inner_text()
                                                exercicio = cols[8].inner_text()
                                                status = cols[9].inner_text()
                                                tipo_processo = cols[10].inner_text()
                                                
                                                vtrDtRecebimento.append(recebimento)
                                                vtrStatus.append(status)

                                                # Converte as datas em objetos datetime
                                                datas_convertidas = [converter_data(data) for data in vtrDtRecebimento]

                                                # Encontra o índice e a data mais atual
                                                i = max(enumerate(datas_convertidas), key=lambda x: x[1])[0]

                                                gravar_informacao(f'dados_pacotes_visualizar_documento_publico_{data_hoje}.txt', f'{cidadeCorrente}, {id_pacote}, {metadado}, {tipo}, {recebimento}, {categoria}, {entidade}, {categoria}, {municipio}, {mes}, {exercicio}, {status}, {tipo_processo}')

                                                atualizar_planilha(usuario, vtrStatus[i], 'G', abaCorrente, linhaCorrente)
                                                atualizar_planilha(usuario, vtrDtRecebimento[i], 'H', abaCorrente, linhaCorrente)
                                                
                                                processar_paginacao_por_numeros(nova_pagina)

                                                time.sleep(2)

                            except Exception as e:
                                gravar_informacao(f'dados_pacotes_visualizar_documento_publico_{data_hoje}.txt', f'{cidadeCorrente}  -  {criterioPesquisa} não encontrado')
                                
                                atualizar_planilha(usuario, f'{criterioPesquisa} não encontrado' , 'G', abaCorrente, linhaCorrente)
                                atualizar_planilha(usuario, '', 'H', abaCorrente, linhaCorrente)

                                print(f"[{datetime.now()}] Erro ao verificar a tabela: {e}")

                            nova_pagina.goto(posicao_dados_url)
                            time.sleep(2)

                # Volta ao link inicial
                nova_pagina.goto(posicao_dados_url)
                print(f"[{datetime.now()}] Todos os índices processados para este município/entidade.")

            # Posição Dados Transmitidos
            else:
                nova_pagina.click('a[href="./iniciarPosicao.do"]')

                time.sleep(2)

                nova_pagina.wait_for_load_state('networkidle')
                posicao_dados_url = nova_pagina.url

                # Espera os filtros estarem visíveis
                if not wait_for_element_with_retry(nova_pagina, 'select[name="mesInicial"]'):
                    print(f"[{datetime.now()}] Erro: Filtro 'Mes Inicial' não carregado.")
                    return

                nova_pagina.select_option('select[name="mesInicial"]', value=mesInicial)
                nova_pagina.select_option('select[name="mesFinal"]', value=mesFinal)
                nova_pagina.select_option('select[name="criterio"]', label=criterioPesquisa)
                nova_pagina.select_option('select[name="anoExercicio"]', value=ano)

                municipio_options = nova_pagina.query_selector_all('select[name="municipio"] option')

                municipio_nome = nova_pagina.locator('select[name="municipio"] option:checked').text_content()
                
                time.sleep(1)

                # Seleção do Município
                if len(municipio_options) > 1:
                    for municipio_index in range(1, len(municipio_options)):  # Começar com o índice 1
                        print(f"[{datetime.now()}] Processando Município {municipio_index}")

                        if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="municipio"]', municipio_index):
                            print(f"[{datetime.now()}] Erro ao selecionar o município {municipio_index}.")
                            continue
                        
                        municipio_nome = nova_pagina.locator('select[name="municipio"] option:checked').text_content()

                        time.sleep(1)

                        # Agora pega todas as opções
                        entidade_options = nova_pagina.query_selector_all('select[name="entidade"] option')

                        print(f"################## Município {municipio_index} Quantidade Total para Entidade {len(entidade_options)}. ##################")

                        if len(entidade_options) > 1:
                            for entidade_index in range(1, len(entidade_options)):  # Começar com o índice 1 para Entidade
                                print(f"[{datetime.now()}] Processando Entidade {entidade_index} para o Município {municipio_index}")
                                
                                if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="entidade"]', entidade_index):
                                    print(f"[{datetime.now()}] Erro ao selecionar a entidade {entidade_index} para o município {municipio_index}.")
                                    continue

                                entidade_nome = nova_pagina.locator('select[name="entidade"] option:checked').text_content()

                                # Verificando e clicando no botão 'Pesquisar'
                                if not wait_for_element_with_retry(nova_pagina, 'img[src="/audesp/img/btn_pesquisar.gif"]:visible'):
                                    print(f"[{datetime.now()}] Erro: Botão 'Pesquisar' não visível.")
                                    return
                                
                                nova_pagina.evaluate('document.querySelector("img[src=\'/audesp/img/btn_pesquisar.gif\']").scrollIntoView()')
                                nova_pagina.click('img[src="/audesp/img/btn_pesquisar.gif"]', force=True)
                                
                                print(f"[{datetime.now()}] Botão 'Pesquisar' clicado para Município {municipio_index}, Entidade {entidade_index}")

                                time.sleep(2)

                                # Verifica se a tabela existe
                                try:
                                    tabela = nova_pagina.query_selector('table')
                                    if tabela:
                                        print(f"[{datetime.now()}] Tabela encontrada para Município {municipio_index}, Entidade {entidade_index}")

                                        table = nova_pagina.query_selector('table.tabelalista')
                                        if table:
                                            print(f"[{datetime.now()}] Tabela encontrada para Entidade {entidade_index}")
                                            
                                            rows = table.query_selector_all('tbody tr')

                                            if len(rows) == 0:
                                                gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')

                                            for row in rows:
                                                cols = row.query_selector_all('td')
                                                if cols:
                                                    id_pacote = cols[0].inner_text()
                                                    remetente = cols[1].inner_text()
                                                    tipo_pacote = cols[2].inner_text()
                                                    data_recepcao = cols[3].inner_text()
                                                    status_pacote = cols[4].inner_text()

                                                    gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}, {id_pacote}, {remetente}, {tipo_pacote}, {data_recepcao}, {status_pacote}')
                                            
                                            processar_paginacao_por_numeros(nova_pagina)

                                            time.sleep(2)
                                    else:
                                        print(f"[{datetime.now()}] Nenhuma tabela encontrada para Município {municipio_index}, Entidade {entidade_index}")

                                        gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')
                                except Exception as e:
                                    print(f"[{datetime.now()}] Erro ao verificar a tabela: {e}")

                                    gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')

                                # Volta para a URL inicial de "Posição Dados Transmitidos"
                                nova_pagina.goto(posicao_dados_url)
                                time.sleep(2)

                        else:
                            #Altera para o primeito Item 
                            nova_pagina.select_option('select[name="entidade"]', index=1)

                            entidade_nome = nova_pagina.locator('select[name="entidade"] option:checked').text_content()
                            
                            # Verificando e clicando no botão 'Pesquisar'
                            if not wait_for_element_with_retry(nova_pagina, 'img[src="/audesp/img/btn_pesquisar.gif"]:visible'):
                                print(f"[{datetime.now()}] Erro: Botão 'Pesquisar' não visível.")
                                return
                            
                            nova_pagina.evaluate('document.querySelector("img[src=\'/audesp/img/btn_pesquisar.gif\']").scrollIntoView()')
                            nova_pagina.click('img[src="/audesp/img/btn_pesquisar.gif"]', force=True)
                            
                            print(f"[{datetime.now()}] Botão 'Pesquisar' clicado para Município {municipio_index}")

                            time.sleep(2)

                            # Verifica se a tabela existe
                            try:
                                tabela = nova_pagina.query_selector('table')
                                if tabela:
                                    print(f"[{datetime.now()}] Tabela encontrada para Município {municipio_index}")

                                    table = nova_pagina.query_selector('table.tabelalista')
                                    if table:
                                        print(f"[{datetime.now()}] Tabela encontrada para Entidade ")
                                        # Lógica para gravar a tabela (se necessário)
            
                                        rows = table.query_selector_all('tbody tr')

                                        if len(rows) == 0:
                                            gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')

                                        for row in rows:
                                            cols = row.query_selector_all('td')
                                            if cols:
                                                id_pacote = cols[0].inner_text()
                                                remetente = cols[1].inner_text()
                                                tipo_pacote = cols[2].inner_text()
                                                data_recepcao = cols[3].inner_text()
                                                status_pacote = cols[4].inner_text()

                                                print('Data Recepção: ', {data_recepcao})

                                                # Aqui você pode substituir pela sua função de gravar as informações
                                                gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}, {id_pacote}, {remetente}, {tipo_pacote}, {data_recepcao}, {status_pacote}')

                                        processar_paginacao_por_numeros(nova_pagina)

                                        time.sleep(2)
                            except Exception as e:
                                print(f"[{datetime.now()}] Erro ao verificar a tabela: {e}")

                                gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')

                            print(f"[{datetime.now()}] Somente uma Entidade para o Município {municipio_index}, pulando para o próximo Município.")
                            nova_pagina.goto(posicao_dados_url)
                            time.sleep(2)

                else:
                    print(f"[{datetime.now()}] Somente um Município, verificando Entidades.")
                    entidade_options = nova_pagina.query_selector_all('select[name="entidade"] option')
                    if len(entidade_options) > 1:
                        for entidade_index in range(1, len(entidade_options)):  # Começar com o índice 1 para Entidade
                            print(f"[{datetime.now()}] Processando Entidade {entidade_index} para o único Município")

                            if not verificar_e_selecionar_combobox(nova_pagina, 'select[name="entidade"]', entidade_index):
                                print(f"[{datetime.now()}] Erro ao selecionar a entidade {entidade_index}.")
                                continue
                            
                            entidade_nome = nova_pagina.locator('select[name="entidade"] option:checked').text_content()
                            
                            # Espera a página carregar e realiza a pesquisa
                            nova_pagina.wait_for_load_state('networkidle', timeout=30000)
                            nova_pagina.click('img[src="/audesp/img/btn_pesquisar.gif"]')
                            print(f"[{datetime.now()}] Botão 'Pesquisar' clicado para Entidade {entidade_index}")

                            # Verifica se a tabela existe
                            try:
                                table = nova_pagina.query_selector('table.tabelalista')
                                if table:
                                    print(f"[{datetime.now()}] Tabela encontrada para Entidade {entidade_index}")
                                    # Lógica para gravar a tabela (se necessário)
        
                                    rows = table.query_selector_all('tbody tr')

                                    if len(rows) == 0:
                                        gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')

                                    for row in rows:
                                        cols = row.query_selector_all('td')
                                        if cols:
                                            id_pacote = cols[0].inner_text()
                                            remetente = cols[1].inner_text()
                                            tipo_pacote = cols[2].inner_text()
                                            data_recepcao = cols[3].inner_text()
                                            status_pacote = cols[4].inner_text()

                                            # Aqui você pode substituir pela sua função de gravar as informações
                                            gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}, {id_pacote}, {remetente}, {tipo_pacote}, {data_recepcao}, {status_pacote}')
                                            
                                    processar_paginacao_por_numeros(nova_pagina)
                                    time.sleep(2)

                                else:
                                    print(f"[{datetime.now()}] Nenhuma tabela encontrada para Entidade {entidade_index}")

                                    gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')

                            except Exception as e:
                                print(f"[{datetime.now()}] Erro ao verificar a tabela: {e}")

                                gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{municipio_nome} - {entidade_nome}  -  {criterioPesquisa} não encontrado')

                            nova_pagina.goto(posicao_dados_url)
                            time.sleep(2)

                # Volta ao link inicial
                nova_pagina.goto(posicao_dados_url)
                print(f"[{datetime.now()}] Todos os índices processados para este município/entidade.")

    except Exception as e:
        atualizar_planilha(usuario, f"[{datetime.now()}] Erro no Playwright: {str(e)}", 'G', linhaCorrente)
        print(f"[{datetime.now()}] Erro no Playwright: {str(e)}")


# Função para converter string em objeto datetime
def converter_data(data_str):
    return datetime.strptime(data_str, "%d/%m/%Y %H:%M:%S")


def processar_paginacao_por_numeros(nova_pagina):
    try:
        # Verifica se a tabela existe
        tabela = nova_pagina.query_selector('table.tabelalista')
        if not tabela:
            print(f"[{datetime.now()}] Nenhuma tabela encontrada.")
            return

        # Verifica se existe a barra de paginação
        pager_footer = nova_pagina.query_selector('tr.pager_footer')
        if not pager_footer:
            print(f"[{datetime.now()}] Nenhuma barra de paginação encontrada.")
            return

        # Coleta os links de página
        page_links = nova_pagina.query_selector_all('tr.pager_footer a')
        
        # Verifica se há mais de uma página e começa na segunda página
        if len(page_links) > 1:
            second_page_link = page_links[1]  # O link da segunda página é o segundo na lista
            print(f"[{datetime.now()}] Avançando para a segunda página: {second_page_link.inner_text()}")
            second_page_link.click()
            
            time.sleep(2)
            # Aguardar até a nova página carregar completamente
            nova_pagina.wait_for_selector('table.tabelalista')  # Aguarda a tabela aparecer na nova página
            tabela = nova_pagina.query_selector('table.tabelalista')  # Re-obtemos a tabela após a navegação

        else:
            print(f"[{datetime.now()}] Apenas uma página, processando os dados.")
            return
        # Laço para processar todas as páginas
        while True:
            # Coleta os dados da tabela atual
        
            rows = tabela.query_selector_all('tbody tr')
            if not rows:
                print(f"[{datetime.now()}] Nenhuma linha encontrada na tabela.")
                break

            for row in rows:
                cols = row.query_selector_all('td')
                if cols:
                    id_pacote = cols[0].inner_text()
                    remetente = cols[1].inner_text()
                    tipo_pacote = cols[2].inner_text()
                    data_recepcao = cols[3].inner_text()
                    status_pacote = cols[4].inner_text()

                    gravar_informacao(f'dados_pacotes_{data_hoje}.txt', f'{id_pacote}, {remetente}, {tipo_pacote}, {data_recepcao}, {status_pacote}')

            # Verifica os links de paginação
            page_links = nova_pagina.query_selector_all('tr.pager_footer a')
            next_page_link = None

            # Encontrar o próximo link da página
            for i in range(len(page_links)):
                link = page_links[i]
                link_text = link.inner_text().strip()
                # Verifica se o link contém o texto de próxima página ou um número não visitado
                if link_text.isdigit():
                    next_page_link = link
                    break

            # Se o link de próxima página foi encontrado, clique e continue
            if next_page_link:
                # Verifica se a próxima página é a mesma que a atual
                next_page_url = next_page_link.get_attribute('href')
                if next_page_url == current_url:
                    print(f"[{datetime.now()}] A página não mudou, encerrando o ciclo.")
                    break

                print(f"[{datetime.now()}] Avançando para a página {next_page_link.inner_text()}")
                next_page_link.click()
                time.sleep(2)

                # Aguardar até a nova página carregar completamente
                nova_pagina.wait_for_selector('table.tabelalista')  # Aguarda a tabela aparecer na nova página

                # Re-obtemos a tabela após a navegação para a próxima página
                tabela = nova_pagina.query_selector('table.tabelalista')
                if not tabela:
                    print(f"[{datetime.now()}] Nenhuma tabela encontrada na página {next_page_link.inner_text()}.")
                    break
                # Atualiza a URL da página atual
                current_url = nova_pagina.url
            else:
                print(f"[{datetime.now()}] Não há mais páginas para processar.")
                break
        
    except Exception as e:
        print(f"[{datetime.now()}] Erro ao processar a paginação por números: {e}")


def verificar_e_selecionar_combobox(pagina, selector, index=1):
    """Função para verificar se o combobox está habilitado, visível e selecionar o índice desejado."""
    try:
        # Aguarda até que o combobox esteja visível e habilitado
        if not wait_for_element_with_retry(pagina, selector):
            print(f"[{datetime.now()}] Combobox {selector} não está visível ou habilitado.")
            return False

        # Forçar o clique no combobox para garantir que a opção seja exibida
        combobox = pagina.query_selector(selector)
        if combobox:
            # Clica no combobox para garantir que ele seja ativado
            combobox.click()

            time.sleep(2)

            print(f"[{datetime.now()}] Combobox {selector} clicado.")

            # Verifica se o índice "0" (Selecione) não está sendo selecionado
            if index > 0:
                # Seleciona o valor do índice no combobox
                pagina.wait_for_load_state('networkidle', timeout=30000)
                pagina.select_option(selector, index=index)
                print(f"[{datetime.now()}] Combobox {selector} selecionado para o índice {index}.")
                return True
            else:
                print(f"[{datetime.now()}] Índice inválido para o combobox {selector}.")
                return False
        else:
            print(f"[{datetime.now()}] Combobox {selector} não encontrado.")
            return False
    except Exception as e:
        print(f"[{datetime.now()}] Erro ao interagir com o combobox {selector}: {e}")
        return False
        

def atualizar_planilha(usuario, valor, coluna, indice_aba = 0, linhaAtual = 0):
    # Obtém o caminho do arquivo
    arquivo = caminho_arquivo_var.get()

    # Verifica se o arquivo existe
    if not os.path.exists(arquivo):
        print("O arquivo especificado não existe.")
        return

    # Verifica se o arquivo tem extensão .xlsx
    if arquivo.endswith(".xlsx"):
        try:
            # Carrega o arquivo Excel com openpyxl
            wb = load_workbook(arquivo)

            # Verifica se o índice da aba é válido
            if indice_aba < 0 or indice_aba >= len(wb.sheetnames):
                print(f"Índice de aba '{indice_aba}' fora do intervalo.")
                return
            
            # sheet = wb.active  # Ativa a planilha atual

            # Carrega a aba especificada pelo índice
            sheet = wb[wb.sheetnames[indice_aba]]  # Usa o índice para acessar a aba

            # Encontra a linha correspondente ao usuário
            usuario_encontrado = False
            for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):  # Assumindo que a 1ª linha é cabeçalho
            # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Assumindo que a 1ª linha é cabeçalho
                # if row[4].value == usuario:  # Supondo que 'Login' esteja na primeira coluna (coluna A)

                 # Garantir que a célula de Login esteja sendo lida corretamente
                usuario_planilha = str(row[4].value).strip().lower()  # Supondo que 'Login' esteja na 5ª coluna (coluna E)

                # if str(row[4].value).strip().lower() == str(usuario).strip().lower():

                if usuario_planilha == str(usuario).strip().lower() and int(index) == int(linhaAtual):  # Comparação normalizada
                    usuario_encontrado = True
                    # Atualiza o valor na coluna especificada
                    row[sheet[coluna + '1'].column - 1].value = valor
                    break

            if not usuario_encontrado:
                print(f"Usuário '{usuario}' não encontrado na planilha. Aba: {indice_aba}")
                return


            # Modifica o valor da célula (linha, coluna são índices baseados em 1)
            sheet.cell(row = 1, column = 7, value = f'Mês de referência: {combobox_mes_referencia.get()}/{ano_combobox.get()}')

            sheet.cell(row = 1, column = 8, value = f'Mês de referência: {combobox_criterio_pesquisa_parametro.get()} Dt. Envio: (Gerado em: {data_hoje})')
            

            # Salva o arquivo Excel com as alterações feitas
            wb.save(arquivo)

            print(f"Planilha atualizada para o usuário {usuario}: {coluna} = {valor}")
        
        except Exception as e:
            print(f"Erro ao atualizar planilha: {e}")
    else:
        print("O arquivo não é um arquivo Excel (.xlsx).")


# Função para validar e formatar a data no formato dd/mm/yyyy
def validar_data(event, campo_data):
    try:
        # Verifica qual campo de data estamos tratando
        if campo_data == "data_inicial":
            texto = entry_data_inicial.get()
        elif campo_data == "data_final":
            texto = entry_data_final.get()

        # Remover caracteres que não são números ou "/"
        texto = re.sub(r'[^0-9/]', '', texto)

        # Caso a string tenha mais de 10 caracteres, limite a 10 (dd/mm/yyyy)
        if len(texto) > 10:
            texto = texto[:10]

        # Aplicar a máscara de data dd/mm/yyyy
        if len(texto) >= 2:
            texto = texto[:2] + '/' + texto[2:]
        if len(texto) >= 5:
            texto = texto[:5] + '/' + texto[5:]

        # Atualizar o conteúdo do Entry com o texto formatado
        if campo_data == "data_inicial":
            entry_data_inicial.delete(0, tk.END)
            entry_data_inicial.insert(0, texto)
        elif campo_data == "data_final":
            entry_data_final.delete(0, tk.END)
            entry_data_final.insert(0, texto)

        # Verificar se a data inserida é válida
        if len(texto) == 10:
            try:
                # Validar a data (se a data for válida, essa linha não gera erro)
                datetime.strptime(texto, "%d/%m/%Y")
                if campo_data == "data_inicial":
                    # Altere o estilo do campo para 'Valid' quando a data for válida
                    entry_data_inicial.config(style="Valid.TEntry")
                elif campo_data == "data_final":
                    # Altere o estilo do campo para 'Valid' quando a data for válida
                    entry_data_final.config(style="Valid.TEntry")
            except ValueError:
                if campo_data == "data_inicial":
                    # Altere o estilo do campo para 'Invalid' quando a data for inválida
                    entry_data_inicial.config(style="Invalid.TEntry")
                elif campo_data == "data_final":
                    # Altere o estilo do campo para 'Invalid' quando a data for inválida
                    entry_data_final.config(style="Invalid.TEntry")
        else:
            if campo_data == "data_inicial":
                entry_data_inicial.config(style="Valid.TEntry")
            elif campo_data == "data_final":
                entry_data_final.config(style="Valid.TEntry")

    except Exception as e:
        print(f"Erro ao validar data: {e}")  # Log do erro para depuração


def aplicar_mascara_data(P):
    """Função para aplicar a máscara de data (dd/mm/yyyy)."""
    # Remover tudo que não for número ou "/"
    P = re.sub(r'[^0-9]', '', P)

    # Limitar a 10 caracteres (dd/mm/yyyy)
    if len(P) > 8:
        P = P[:8]

    # Adicionar as barras conforme a máscara de data
    if len(P) >= 2:
        P = P[:2] + '/' + P[2:]
    if len(P) >= 5:
        P = P[:5] + '/' + P[5:]

    return P


def on_keyrelease(event, entry):
    """Função que aplica a máscara enquanto o usuário digita no Entry."""
    texto = entry.get()
    texto_formatado = aplicar_mascara_data(texto)
    entry.delete(0, tk.END)
    entry.insert(0, texto_formatado)


def atualizar_criterios(event=None):
    # Pega o valor selecionado no primeiro combobox
    documento_selecionado = combobox_documento_parametro.get()

    # Verifica o valor selecionado no primeiro combobox e ajusta os valores do segundo
    if documento_selecionado == "Visualizar Documentos Públicos":
        novos_valores =  ["ATA-AUDIENCIA-ACOES-SAUDE", "ATA-AUDIENCIA-AVALIAÇÃO-CUMPRIMENTO-METAS", "ATA-REUNIÃO-CONSELHO-FUNDEF", "Ato de fixação da remuneração de agentes políticos"
                        , "Atualização do Cadastro Geral de Entidades", "Atualização do Cadastro Geral de Entidades - Mensal", "BALANCETE-CONJUNTO-CONTA-CONTABIL", "BALANCETE-CONJUNTO-CONTA-CORRENTE"
                        , "BALANCETE-CONJUNTO-ENCERRAMENTO-13-CONTA-CONTABIL", "BALANCETE-CONJUNTO-ENCERRAMENTO-13-CONTA-CORRENTE", "BALANCETE-CONJUNTO-ENCERRAMENTO-14-CONTA-CONTABIL", "BALANCETE-CONJUNTO-ENCERRAMENTO-14-CONTA-CORRENTE"
                        , "BALANCETE-CONJUNTO-ENCERRAMENTO-FINAL-CONTA-CONTABIL", "BALANCETE-CONJUNTO-ENCERRAMENTO-FINAL-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-CONTA-CONTABIL", "BALANCETE-CONSOLIDADO-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-13-CONTA-CONTABIL"
                        , "BALANCETE-CONSOLIDADO-ENCERRAMENTO-13-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-14-CONTA-CONTABIL", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-14-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-FINAL-CONTA-CONTABIL"
                        , "BALANCETE-CONSOLIDADO-ENCERRAMENTO-FINAL-CONTA-CORRENTE", "BALANCETE-ISOLADO-CONTA-CONTABIL", "BALANCETE-ISOLADO-CONTA-CORRENTE", "BALANCETE-ISOLADO-ENCERRAMENTO-13-CONTA-CONTABIL", "BALANCETE-ISOLADO-ENCERRAMENTO-13-CONTA-CORRENTE"
                        , "BALANCETE-ISOLADO-ENCERRAMENTO-14-CONTA-CONTABIL", "BALANCETE-ISOLADO-ENCERRAMENTO-14-CONTA-CORRENTE", "BALANCETE-ISOLADO-ENCERRAMENTO-FINAL-CONTA-CONTABIL", "BALANCETE-ISOLADO-ENCERRAMENTO-FINAL-CONTA-CORRENTE", "Cadastro de Fundos de Investimento"
                        , "Cadastro de Parcelamentos com RPPS", "Cadastro Eletrônico de Obras em Execução", "Cadastro Eletrônico de Obras em Execução - 2012", "CADASTRO-PLANEJAMENTO", "CADASTRO-PLANEJAMENTO-ATUALIZADO-2010", "CADASTROS-CONTABEIS", "Complemento de Conciliações Bancárias"
                        , "Complemento de Relatório de Atividades", "Complemento de Remuneração de Agentes Políticos", "Concessão de Reajuste de Agentes Políticos", "Conciliações Bancárias", "Conciliações Bancárias Mensais", "Dados de Balanços Conjuntos", "Dados de Balanços Isolados"
                        , "DECRETO-REGULAMENTAÇÃO-FUNDO-SAÚDE", "Demonstrativo da Rentabilidade e Evolução dos Investimentos", "Demonstrativo de Receitas Previdenciárias", "ESTATUTO-REGIMENTO-REGULAMENTAÇÃO", "Fixação da Remuneração de Agentes Políticos", "Fundos de Investimento"
                        , "LDO-ALTERACAO-ATA-AUDIENCIA-APROVAÇÃO", "LDO-ALTERACAO-ATA-AUDIENCIA-ELABORACAO", "LDO-ATUALIZACAO", "LDO-ATUALIZADA-2010", "LDO-INICIAL", "LDO-INICIAL-ATA-AUDIENCIA-APROVAÇÃO", "LDO-INICIAL-ATA-AUDIENCIA-ELABORACAO", "LDO-LEI-ATUALIZACAO", "LDO-LEI-INICIAL"
                        , "LEI-ADIANTAMENTO", "LEI-CRIAÇÃO-CONSELHO-EDUCAÇÃO", "LEI-CRIAÇÃO-CONSELHO-SAÚDE", "LEI-CRIAÇÃO-FUNDO-SAUDE", "LEI-CRIAÇÃO-INSTITUIÇÃO", "LEI-ORGANICA", "LOA-ALTERACAO-ATA-AUDIENCIA-APROVAÇÃO", "LOA-ALTERACAO-ATA-AUDIENCIA-ELABORACAO", "LOA-ATUALIZACAO"
                        , "LOA-ATUALIZADA-2010", "LOA-INICIAL", "LOA-INICIAL-ATA-AUDIENCIA-APROVAÇÃO", "LOA-INICIAL-ATA-AUDIENCIA-ELABORACAO", "LOA-LEI-ATUALIZACAO", "LOA-LEI-INICIAL", "Mapa de Precatórios", "NORMA-INSTITUIÇÃO-CONSELHO-FUNDEF-FUNDEB", "Parcelamentos com RPPS"
                        , "PARECER-CONSELHO-FUNDEB", "PARECER-CONSELHO-SAUDE", "PECAS-PLANEJAMENTO", "PLAN-CADASTRO", "PLAN-CADASTRO-ODS", "PLAN-LDO-ATUALIZADA", "PLAN-LDO-INICIAL", "PLAN-LOA-ATUALIZADA", "PLAN-LOA-INICIAL", "PLANO-CARREIRA-MAGISTÉRIO", "PLANO-MUNICIPAL-AÇÕES ANUAIS-SAUDE"
                        , "PLAN-PPA-ATUALIZADO", "PLAN-PPA-INICIAL", "PPA-ALTERACAO-ATA-AUDIENCIA-APROVAÇÃO", "PPA-ALTERACAO-ATA-AUDIENCIA-ELABORACAO", "PPA-ATUALIZACAO", "PPA-ATUALIZADO-2010", "PPA-INICIAL", "PPA-INICIAL-ATA-AUDIENCIA-APROVAÇÃO", "PPA-INICIAL-ATA-AUDIENCIA-ELABORACAO"
                        , "PPA-LEI-ATUALIZACAO", "PPA-LEI-INICIAL", "Publ. Aplic. na Manut. e Desenv. do Ensino", "Publ. Demonstrativo de Projeção Atuarial do RPPS", "Publ. do Demonst. de Receitas e Despesas com Manutenção e Desenvolvimento do Ensino (Anexo 8 RREO)"
                        , "Publ. do Demonstrativo das Receitas e Despesas com Ações e Serviços Públicos  de Saúde", "Publ. Remuneração Cargos e Empregos Públicos", "Publ. RGF - Executivo", "Publ. RGF - Legislativo", "Publ. RREO - Aplic. Recursos de Alienação de Ativos"
                        , "Publ. RREO - Balanço Orçamentário", "Publ. RREO - Dem. Apuração RCL", "Publ. RREO - Dem. Função / Subfunção", "Publ. RREO - Dem. Receitas e Despesas Previdenciárias", "Publ. RREO - Oper. Crédito X Desp. Capital", "Publ. RREO - Projeção Atuarial do RPPS"
                        , "Publ. RREO - Restos a Pagar", "Publ. RREO - Resultado Nominal", "Publ. RREO - Resultado Primário", "Questionário de Contratos de Programa", "Questionário de Serviços de Saneamento Básico", "Questionário sobre Quadro de Pessoal (a partir de 2016)"
                        , "Questionário sobre Quadro de Pessoal e Transporte (somente 2015)", "Recibo de Prestação de Contas", "REGIMENTO-INTERNO-CONSELHO-SAÚDE", "Relação de Contratos de Concessão e Permissão de Serviço Público", "RELATÓRIO DE ALERTA", "RELATÓRIO DE ALERTA - SUBSTITUÍDO"
                        , "Relatório de Atividades", "RELATÓRIO DE INCONSISTÊNCIA", "RELATÓRIO DE INSTRUÇÃO", "RELATÓRIO DE INSTRUÇÃO - SUBSTITUÍDO", "Remuneração de Agentes Políticos", "SIAP", "SisCAA", "SisRTS", "TABELASAUXILIARES", "TERMO-CONVENIO-MUNICIPALIZAÇÃO-ENSINO"
                        , "Termo de consentimento de acesso a contas bancárias", "TIPOSGENERICOS"]
        
        # Definindo valores iniciais para as datas
        hoje = datetime.now()
        entry_data_inicial.insert(0, f"01/01/{hoje.year}")
        entry_data_final.insert(0, f"{hoje.day:02}/{hoje.month:02}/{hoje.year}")

        # Mostrar os campos de data
        label_data_inicial.grid(row=8, column=0, padx=5, pady=5, sticky="w")
        entry_data_inicial.grid(row=8, column=1, padx=5, pady=5, sticky="w")
        label_data_final.grid(row=9, column=0, padx=5, pady=5, sticky="w")
        entry_data_final.grid(row=9, column=1, padx=5, pady=5, sticky="w")

        # Habilitar os campos de data
        entry_data_inicial.config(state="normal")
        entry_data_final.config(state="normal")
       
        combobox_mes_inicial.grid_forget()
        combobox_mes_fim.grid_forget()
        
        label_mes_inicial.grid_forget()
        label_mes_fim.grid_forget()

        label_mes_referencia.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        combobox_mes_referencia.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        label_combobox_tipo_processo.grid(row=10, column=0, padx=5, pady=5, sticky="w")
        combobox_tipo_processo.grid(row=10, column=1, padx=5, pady=5, sticky="w")

        button_processar.grid(row=12, column=1, padx=5, pady=5, sticky="w")

        # Criando o Checkbutton ao lado do botão
        checkbox_background.grid(row=12, column=1, padx=110, pady=5, sticky="ew")


        label_combobox_entidades.grid(row=11, column=0, padx=5, pady=5, sticky="w")
        combobox_entidades.grid(row=11, column=1, padx=5, pady=5, sticky="w")
    

        combobox_criterio_pesquisa_parametro.set('BALANCETE-ISOLADO-CONTA-CORRENTE')  # Limpa a seleção
    
    elif documento_selecionado == "Posição Dados Transmitidos":
        novos_valores = ["ATAS-PARECERES-CONSELHOS", "AUDIÊNCIAS-PÚBLICAS-ACOMPANHAMENTO", "AUDIÊNCIAS-PÚBLICAS-APROVAÇÃO", 
            "AUDIÊNCIAS-PÚBLICAS-ELABORAÇÃO", "BALANCETE-CONJUNTO", "BALANCETE-CONJUNTO-ENCERRAMENTO",
            "BALANCETE-CONJUNTO-ENCERRAMENTO-13", "BALANCETE-CONJUNTO-ENCERRAMENTO-14", "BALANCETE-CONSOLIDADO", 
            "BALANCETE-CONSOLIDADO-ENCERRAMENTO", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-13", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-14",
            "BALANCETE-ISOLADO", "BALANCETE-ISOLADO-ENCERRAMENTO", "BALANCETE-ISOLADO-ENCERRAMENTO-13", "BALANCETE-ISOLADO-ENCERRAMENTO-14", 
            "BALANÇOS CONJUNTOS", "BALANÇOS CONSOLIDADOS", "BALANÇOS ISOLADOS", "CADASTRO ELETRÔNICO DE OBRAS EM EXECUÇÃO", 
            "CADASTROS-CONTABEIS", "CONCESSÃO DE REAJUSTE DA REMUNERAÇÃO DE AGENTES POLÍTICOS", "CONCILIAÇÕES BANCÁRIAS", 
            "Demonstrativo da Rentabilidade e Evolução dos Investimentos", "Demonstrativo de Receitas Previdenciárias",
            "FIXAÇÃO DA REMUNERAÇÃO DE AGENTES POLÍTICOS", "FUNDOS DE INVESTIMENTO", "LEI-ADIANTAMENTO", "LEI-CRIAÇÃO-REGULAMENTO-ENTIDADE",
            "LEIS-NORMAS-PLANO-ENSINO", "LEIS-NORMAS-PLANO-SAÚDE", "MAPA DE PRECATÓRIOS", "Parcelamentos com RPPS", 
            "PLANEJAMENTO-ATUALIZACAO-2009", "PLANEJAMENTO-ATUALIZADO", "PLANEJAMENTO-ATUALIZADO-2010", "PLANEJAMENTO-CADASTRO", 
            "PLANEJAMENTO-CADASTRO-ODS", "PLANEJAMENTO-INICIAL", "PLANEJAMENTO-INICIAL-2009", "PLANEJAMENTO-INICIAL-2010", 
            "PLANEJAMENTO-LEIS", "PLANEJAMENTO-PPA-INICIAL", "RELAÇÃO DE CONTRATOS DE CONCESSÃO", "RELATÓRIO DE ATIVIDADES", 
            "REMUNERAÇÃO DE AGENTES POLÍTICOS", "SIAP", "Termo de consentimento de acesso a contas bancárias"]
        
        # Esconder os campos de data
        label_data_inicial.grid_forget()
        entry_data_inicial.grid_forget()
        label_data_final.grid_forget()
        entry_data_final.grid_forget()
        
        # Desabilitar os campos de data
        entry_data_inicial.config(state="disabled")
        entry_data_final.config(state="disabled")

        label_mes_inicial.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        label_mes_fim.grid(row=4, column=0, padx=5, pady=5, sticky="w")

        combobox_mes_inicial.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        combobox_mes_fim.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        label_mes_referencia.grid_forget()
        combobox_mes_referencia.grid_forget()

        label_combobox_tipo_processo.grid_forget()
        combobox_tipo_processo.grid_forget()

        # Botão para processar (movido para a linha 8)
        button_processar.grid(row=9, column=1, padx=5, pady=5, sticky="w")

        # Criando o Checkbutton ao lado do botão
        checkbox_background.grid(row=9, column=1, padx=110, pady=5, sticky="ew")

        label_combobox_entidades.grid(row=8, column=0, padx=5, pady=5, sticky="w")
        combobox_entidades.grid(row=8, column=1, padx=5, pady=5, sticky="w")

        combobox_criterio_pesquisa_parametro.set('')  # Limpa a seleção
        combobox_criterio_pesquisa_parametro.set(novos_valores[8])
    else:
        novos_valores = []

    # Atualiza os valores do segundo combobox
    combobox_criterio_pesquisa_parametro['values'] = novos_valores
    # combobox_criterio_pesquisa_parametro.set('')  # Limpa a seleção
    # combobox_criterio_pesquisa_parametro.set(novos_valores[0])

# Cria a janela principal
root = tk.Tk()
root.title("Selecionar e Processar Arquivo")

# Bloqueia o redimensionamento da janela
root.resizable(False, False)

# Variáveis para armazenar os parâmetros
caminho_arquivo_var = tk.StringVar()
link_var = tk.StringVar()  # Variável para o link informado na tela
ano_var = tk.StringVar(value=str(datetime.now().year))  # Ano com valor padrão
mes_inicial_var = tk.StringVar(value="1")  # Mês com valor padrão (1)
mes_final_var = tk.StringVar(value="1")  # Mês com valor padrão (1)
combobox_menu_var = tk.StringVar(value="Audesp")  # Combobox com valor padrão (default01)
combobox_perfil_var = tk.StringVar(value="Audesp Base - Prestação de dados")  # Combobox com valor padrão (default01)
combobox_criterio_pesquisa_var = tk.StringVar(value="BALANCETE-ISOLADO-CONTA-CORRENTE")  # Combobox com valor padrão (default01)

combobox_documento_var = tk.StringVar(value="Visualizar Documentos Públicos")  # Combobox com valor padrão (default01)
combobox_documento_var.set("Visualizar Documentos Públicos")

combobox_criterio_pesquisa_var.set("BALANCETE-ISOLADO-CONTA-CORRENTE") 

mes_referencia_var = tk.StringVar(value="1")  # Mês com valor padrão (1)
link_var= tk.StringVar(value="https://sso.tce.sp.gov.br/cas-server/login?service=https%3A%2F%2Fsso.tce.sp.gov.br%2FPortal/j_spring_cas_security_check")


# Layout da interface utilizando grid
tk.Label(root, text="Arquivo:").grid(row=0, column=0, padx=5, pady=5, sticky="w")

# Entrada de arquivo (somente leitura)
entrada_arquivo = tk.Entry(root, textvariable=caminho_arquivo_var, width=70, state="readonly")
entrada_arquivo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

button_selecionar_arquivo = tk.Button(root, text="Selecionar Arquivo", command=selecionar_arquivo)
button_selecionar_arquivo.grid(row=0, column=2, padx=5, pady=5, sticky="w")

# Parâmetro de link (campo de texto para informar um link)
tk.Label(root, text="URL:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entrada_url = tk.Entry(root, textvariable=link_var, width=70)
entrada_url.grid(row=1, column=1, padx=5, pady=5, sticky="w")

# Parâmetro de ano (Combobox com o ano atual)
tk.Label(root, text="Ano:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
ano_combobox = ttk.Combobox(root, textvariable=ano_var, values=[str(year) for year in range(2000, datetime.now().year + 1)], state="readonly")
ano_combobox.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# Parâmetro de mês (Combobox de 1 a 12) - Mês Início
label_mes_inicial = tk.Label(root, text="Mês Início:")
label_mes_inicial.grid(row=3, column=0, padx=5, pady=5, sticky="w")
combobox_mes_inicial = ttk.Combobox(root, textvariable=mes_inicial_var, values=[str(i) for i in range(1, 13)], state="readonly", width=10)
combobox_mes_inicial.grid(row=3, column=1, padx=5, pady=5, sticky="w")

# Parâmetro de mês (Combobox de 1 a 12) - Mês Fim
label_mes_fim = tk.Label(root, text="Mês Fim:")
label_mes_fim.grid(row=4, column=0, padx=5, pady=5, sticky="w")
combobox_mes_fim = ttk.Combobox(root, textvariable=mes_final_var, values=[str(i) for i in range(1, 13)], state="readonly", width=10)
combobox_mes_fim.grid(row=4, column=1, padx=5, pady=5, sticky="w")

combobox_tipo_processo_var = tk.StringVar(value="Selecione...")  # Mês com valor padrão (1)
checkbox_background_var = tk.IntVar(value=1)

combobox_entidades_var = tk.StringVar(value="Todos")  

# Parâmetro de critério de pesquisa (Combobox com valores padrão)
tk.Label(root, text="Critério de Pesquisa:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
combobox_criterio_pesquisa_parametro = ttk.Combobox(root, textvariable=combobox_criterio_pesquisa_var,
    values=["ATA-AUDIENCIA-ACOES-SAUDE", "ATA-AUDIENCIA-AVALIAÇÃO-CUMPRIMENTO-METAS", "ATA-REUNIÃO-CONSELHO-FUNDEF", "Ato de fixação da remuneração de agentes políticos"
                        , "Atualização do Cadastro Geral de Entidades", "Atualização do Cadastro Geral de Entidades - Mensal", "BALANCETE-CONJUNTO-CONTA-CONTABIL", "BALANCETE-CONJUNTO-CONTA-CORRENTE"
                        , "BALANCETE-CONJUNTO-ENCERRAMENTO-13-CONTA-CONTABIL", "BALANCETE-CONJUNTO-ENCERRAMENTO-13-CONTA-CORRENTE", "BALANCETE-CONJUNTO-ENCERRAMENTO-14-CONTA-CONTABIL", "BALANCETE-CONJUNTO-ENCERRAMENTO-14-CONTA-CORRENTE"
                        , "BALANCETE-CONJUNTO-ENCERRAMENTO-FINAL-CONTA-CONTABIL", "BALANCETE-CONJUNTO-ENCERRAMENTO-FINAL-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-CONTA-CONTABIL", "BALANCETE-CONSOLIDADO-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-13-CONTA-CONTABIL"
                        , "BALANCETE-CONSOLIDADO-ENCERRAMENTO-13-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-14-CONTA-CONTABIL", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-14-CONTA-CORRENTE", "BALANCETE-CONSOLIDADO-ENCERRAMENTO-FINAL-CONTA-CONTABIL"
                        , "BALANCETE-CONSOLIDADO-ENCERRAMENTO-FINAL-CONTA-CORRENTE", "BALANCETE-ISOLADO-CONTA-CONTABIL", "BALANCETE-ISOLADO-CONTA-CORRENTE", "BALANCETE-ISOLADO-ENCERRAMENTO-13-CONTA-CONTABIL", "BALANCETE-ISOLADO-ENCERRAMENTO-13-CONTA-CORRENTE"
                        , "BALANCETE-ISOLADO-ENCERRAMENTO-14-CONTA-CONTABIL", "BALANCETE-ISOLADO-ENCERRAMENTO-14-CONTA-CORRENTE", "BALANCETE-ISOLADO-ENCERRAMENTO-FINAL-CONTA-CONTABIL", "BALANCETE-ISOLADO-ENCERRAMENTO-FINAL-CONTA-CORRENTE", "Cadastro de Fundos de Investimento"
                        , "Cadastro de Parcelamentos com RPPS", "Cadastro Eletrônico de Obras em Execução", "Cadastro Eletrônico de Obras em Execução - 2012", "CADASTRO-PLANEJAMENTO", "CADASTRO-PLANEJAMENTO-ATUALIZADO-2010", "CADASTROS-CONTABEIS", "Complemento de Conciliações Bancárias"
                        , "Complemento de Relatório de Atividades", "Complemento de Remuneração de Agentes Políticos", "Concessão de Reajuste de Agentes Políticos", "Conciliações Bancárias", "Conciliações Bancárias Mensais", "Dados de Balanços Conjuntos", "Dados de Balanços Isolados"
                        , "DECRETO-REGULAMENTAÇÃO-FUNDO-SAÚDE", "Demonstrativo da Rentabilidade e Evolução dos Investimentos", "Demonstrativo de Receitas Previdenciárias", "ESTATUTO-REGIMENTO-REGULAMENTAÇÃO", "Fixação da Remuneração de Agentes Políticos", "Fundos de Investimento"
                        , "LDO-ALTERACAO-ATA-AUDIENCIA-APROVAÇÃO", "LDO-ALTERACAO-ATA-AUDIENCIA-ELABORACAO", "LDO-ATUALIZACAO", "LDO-ATUALIZADA-2010", "LDO-INICIAL", "LDO-INICIAL-ATA-AUDIENCIA-APROVAÇÃO", "LDO-INICIAL-ATA-AUDIENCIA-ELABORACAO", "LDO-LEI-ATUALIZACAO", "LDO-LEI-INICIAL"
                        , "LEI-ADIANTAMENTO", "LEI-CRIAÇÃO-CONSELHO-EDUCAÇÃO", "LEI-CRIAÇÃO-CONSELHO-SAÚDE", "LEI-CRIAÇÃO-FUNDO-SAUDE", "LEI-CRIAÇÃO-INSTITUIÇÃO", "LEI-ORGANICA", "LOA-ALTERACAO-ATA-AUDIENCIA-APROVAÇÃO", "LOA-ALTERACAO-ATA-AUDIENCIA-ELABORACAO", "LOA-ATUALIZACAO"
                        , "LOA-ATUALIZADA-2010", "LOA-INICIAL", "LOA-INICIAL-ATA-AUDIENCIA-APROVAÇÃO", "LOA-INICIAL-ATA-AUDIENCIA-ELABORACAO", "LOA-LEI-ATUALIZACAO", "LOA-LEI-INICIAL", "Mapa de Precatórios", "NORMA-INSTITUIÇÃO-CONSELHO-FUNDEF-FUNDEB", "Parcelamentos com RPPS"
                        , "PARECER-CONSELHO-FUNDEB", "PARECER-CONSELHO-SAUDE", "PECAS-PLANEJAMENTO", "PLAN-CADASTRO", "PLAN-CADASTRO-ODS", "PLAN-LDO-ATUALIZADA", "PLAN-LDO-INICIAL", "PLAN-LOA-ATUALIZADA", "PLAN-LOA-INICIAL", "PLANO-CARREIRA-MAGISTÉRIO", "PLANO-MUNICIPAL-AÇÕES ANUAIS-SAUDE"
                        , "PLAN-PPA-ATUALIZADO", "PLAN-PPA-INICIAL", "PPA-ALTERACAO-ATA-AUDIENCIA-APROVAÇÃO", "PPA-ALTERACAO-ATA-AUDIENCIA-ELABORACAO", "PPA-ATUALIZACAO", "PPA-ATUALIZADO-2010", "PPA-INICIAL", "PPA-INICIAL-ATA-AUDIENCIA-APROVAÇÃO", "PPA-INICIAL-ATA-AUDIENCIA-ELABORACAO"
                        , "PPA-LEI-ATUALIZACAO", "PPA-LEI-INICIAL", "Publ. Aplic. na Manut. e Desenv. do Ensino", "Publ. Demonstrativo de Projeção Atuarial do RPPS", "Publ. do Demonst. de Receitas e Despesas com Manutenção e Desenvolvimento do Ensino (Anexo 8 RREO)"
                        , "Publ. do Demonstrativo das Receitas e Despesas com Ações e Serviços Públicos  de Saúde", "Publ. Remuneração Cargos e Empregos Públicos", "Publ. RGF - Executivo", "Publ. RGF - Legislativo", "Publ. RREO - Aplic. Recursos de Alienação de Ativos"
                        , "Publ. RREO - Balanço Orçamentário", "Publ. RREO - Dem. Apuração RCL", "Publ. RREO - Dem. Função / Subfunção", "Publ. RREO - Dem. Receitas e Despesas Previdenciárias", "Publ. RREO - Oper. Crédito X Desp. Capital", "Publ. RREO - Projeção Atuarial do RPPS"
                        , "Publ. RREO - Restos a Pagar", "Publ. RREO - Resultado Nominal", "Publ. RREO - Resultado Primário", "Questionário de Contratos de Programa", "Questionário de Serviços de Saneamento Básico", "Questionário sobre Quadro de Pessoal (a partir de 2016)"
                        , "Questionário sobre Quadro de Pessoal e Transporte (somente 2015)", "Recibo de Prestação de Contas", "REGIMENTO-INTERNO-CONSELHO-SAÚDE", "Relação de Contratos de Concessão e Permissão de Serviço Público", "RELATÓRIO DE ALERTA", "RELATÓRIO DE ALERTA - SUBSTITUÍDO"
                        , "Relatório de Atividades", "RELATÓRIO DE INCONSISTÊNCIA", "RELATÓRIO DE INSTRUÇÃO", "RELATÓRIO DE INSTRUÇÃO - SUBSTITUÍDO", "Remuneração de Agentes Políticos", "SIAP", "SisCAA", "SisRTS", "TABELASAUXILIARES", "TERMO-CONVENIO-MUNICIPALIZAÇÃO-ENSINO"
                        , "Termo de consentimento de acesso a contas bancárias", "TIPOSGENERICOS"], state="readonly", width=70)
combobox_criterio_pesquisa_parametro.grid(row=5, column=1, padx=5, pady=5, sticky="w")

# Parâmetro de combobox com valores padrão
label_menu_parametro = tk.Label(root, text="Menu:")
label_menu_parametro.grid(row=6, column=0, padx=5, pady=5, sticky="w")

combobox_menu_parametro = ttk.Combobox(root, textvariable=combobox_menu_var, values=["default01", "default02"], state="readonly", width=70)
combobox_menu_parametro.grid(row=6, column=1, padx=5, pady=5, sticky="w")
combobox_menu_parametro.config(state="disabled")

# Parâmetro de combobox com valores padrão
tk.Label(root, text="Perfil:").grid(row=7, column=0, padx=5, pady=5, sticky="w")
combobox_perfil_parametro = ttk.Combobox(root, textvariable=combobox_perfil_var, values=["default01", "default02"], state="readonly", width=70)
combobox_perfil_parametro.grid(row=7, column=1, padx=5, pady=5, sticky="w")

# Parâmetro de combobox com valores padrão
tk.Label(root, text="Documento:").grid(row=7, column=0, padx=5, pady=5, sticky="w")
combobox_documento_parametro = ttk.Combobox(root, textvariable=combobox_documento_var, values=["Visualizar Documentos Públicos", "Posição Dados Transmitidos"], state="readonly", width=70)
combobox_documento_parametro.grid(row=7, column=1, padx=5, pady=5, sticky="w")

# Parâmetro de mês (Combobox de 1 a 12) - Mês Referência
label_mes_referencia = tk.Label(root, text="Mês Referência:")
label_mes_referencia.grid_forget()
combobox_mes_referencia = ttk.Combobox(root, textvariable=mes_referencia_var, values=["Selecione..."] + [str(i) for i in range(1, 13)], state="readonly", width=10)
combobox_mes_referencia.grid(row=3, column=1, padx=5, pady=5, sticky="w")

# Campos de Data (inicialmente invisíveis)
label_data_inicial = tk.Label(root, text="Data Inicial Recebimento:")
entry_data_inicial = ttk.Entry(root, width=15)
label_data_final = tk.Label(root, text="Data Final Recebimento:")
entry_data_final = ttk.Entry(root, width=15)





# Parâmetro de critério de pesquisa (Combobox com valores padrão)
label_combobox_tipo_processo = tk.Label(root, text="Tipo de Processo:")
label_combobox_tipo_processo.grid_forget()
combobox_tipo_processo = ttk.Combobox(root, textvariable=combobox_tipo_processo_var,
    values=["Selecione...", "ACOMPANHAMENTO DA GESTÃO FISCAL", "ADMISSÃO - CONCURSO / PROCESSO SELETIVO", "ADMISSÃO - TEMPO DETERMINADO"
          , "APOSENTADORIA", "AUXÍLIOS/SUBVENÇÕES/CONTRIBUIÇÕES", "CONTAS ANUAIS", "CONTRATO", "PENSÃO", "RPPS"], state="readonly", width=70)
combobox_tipo_processo.grid_forget()

# Função de validação de tecla para a máscara de data
def validar_tecla(event):
    """Validar tecla pressionada para máscara de data."""
    texto = event.widget.get()

    # Aplicar a máscara
    texto_mascarado = aplicar_mascara_data(texto)

    # Atualizar o campo com a máscara
    event.widget.delete(0, tk.END)
    event.widget.insert(0, texto_mascarado)

# Função para rodar o Playwright em uma thread separada
def start_process():
    # Usar threading para rodar o Playwright sem travar o Tkinter
    threading.Thread(target=processar_arquivo, daemon=True).start()


# Vincula a função de validação de tecla para os campos de data
entry_data_inicial.bind("<KeyRelease>", validar_tecla)
entry_data_final.bind("<KeyRelease>", validar_tecla)

# Botão para processar 
button_processar = tk.Button(root, text="Processar Arquivo", command=start_process)
button_processar.grid(row=8, column=1, padx=5, pady=5, sticky="w")


# Criando o Checkbutton ao lado do botão
checkbox_background = tk.Checkbutton(root, width="20", text="Executar em segundo plano", variable=checkbox_background_var)
checkbox_background.grid(row=8, column=1, padx=100, pady=5, sticky="ew")

# Vincula a função que irá atualizar o segundo combobox com a mudança do primeiro
combobox_documento_parametro.bind("<<ComboboxSelected>>", atualizar_criterios)

# Obter a data de hoje no formato dd/mm/yyyy
data_hoje = datetime.now().strftime('%d-%m-%Y')



# Variável global para armazenar a barra de progresso
progress_bar = None

# Cria uma nova barra de progresso com o novo comprimento
# progress_bar = ttk.Progressbar(root, orient="horizontal", length=100, mode="determinate")
# progress_bar.grid_forget()


label_progresso = tk.Label(root, text="")
label_progresso.grid_forget()



# 


# Parâmetro de combobox com valores padrão
label_combobox_entidades = tk.Label(root, text="Entidade(s):")
label_combobox_entidades.grid(row=11, column=0, padx=5, pady=5, sticky="w")
combobox_entidades = ttk.Combobox(root, textvariable=combobox_entidades_var, values=["Todos", "Prefeitura", "Câmara", "Outros"], state="readonly", width=70)
combobox_entidades.grid(row=11, column=1, padx=5, pady=5, sticky="w")


atualizar_criterios(None)  # Chama a função imediatamente após a definição do valor
# Inicia o loop principal da interface gráfica
root.mainloop()
